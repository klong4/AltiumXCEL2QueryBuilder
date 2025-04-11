#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel Formatter
=============

Provides advanced Excel formatting capabilities for pivot tables and rule data.
"""

import logging
import os
from typing import Dict, List, Optional, Union, Tuple, Any

import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.styles.colors import Color
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart, ScatterChart
from openpyxl.chart.series import Series

from models.rule_model import UnitType, RuleType

logger = logging.getLogger(__name__)

class ExcelFormattingPreset:
    """Represents a preset for Excel formatting"""

    def __init__(self, name: str, description: str):
        """Initialize a formatting preset"""
        self.name = name
        self.description = description
        self.header_font = None
        self.header_fill = None
        self.data_font = None
        self.data_fill = None
        self.border = None
        self.alignment = None
        self.conditional_formatting = []
        self.chart_type = None
        self.include_summary = False
        self.include_statistics = False
        self.freeze_panes = True
        self.auto_filter = True

    def to_dict(self) -> Dict:
        """Convert preset to dictionary for serialization"""
        return {
            "name": self.name,
            "description": self.description,
            "include_summary": self.include_summary,
            "include_statistics": self.include_statistics,
            "freeze_panes": self.freeze_panes,
            "auto_filter": self.auto_filter,
            "chart_type": self.chart_type
        }
    
    @classmethod
    def from_dict(cls, data: Dict) -> 'ExcelFormattingPreset':
        """Create preset from dictionary"""
        preset = cls(
            name=data.get("name", "Custom Preset"),
            description=data.get("description", "")
        )
        preset.include_summary = data.get("include_summary", False)
        preset.include_statistics = data.get("include_statistics", False)
        preset.freeze_panes = data.get("freeze_panes", True)
        preset.auto_filter = data.get("auto_filter", True)
        preset.chart_type = data.get("chart_type", None)
        return preset


class ExcelFormatter:
    """Handles advanced Excel formatting for exports"""

    # Predefined presets
    PRESETS = {
        "default": {
            "name": "Default",
            "description": "Simple, clean formatting for pivot tables",
            "include_summary": False,
            "include_statistics": False,
            "freeze_panes": True,
            "auto_filter": True,
            "chart_type": None
        },
        "professional": {
            "name": "Professional",
            "description": "Polished formatting suitable for reports",
            "include_summary": True,
            "include_statistics": True,
            "freeze_panes": True,
            "auto_filter": True,
            "chart_type": "bar"
        },
        "technical": {
            "name": "Technical",
            "description": "Detailed formatting with statistics and visualization",
            "include_summary": True,
            "include_statistics": True,
            "freeze_panes": True,
            "auto_filter": True,
            "chart_type": "heatmap"
        },
        "minimal": {
            "name": "Minimal",
            "description": "Basic formatting with no extras",
            "include_summary": False,
            "include_statistics": False,
            "freeze_panes": False,
            "auto_filter": False,
            "chart_type": None
        }
    }

    def __init__(self):
        """Initialize Excel formatter"""
        self.current_preset = self.get_preset("default")
        logger.info("Excel formatter initialized with default preset")

    def get_preset(self, preset_id: str) -> ExcelFormattingPreset:
        """Get a formatting preset by ID"""
        if preset_id in self.PRESETS:
            return ExcelFormattingPreset.from_dict(self.PRESETS[preset_id])
        logger.warning(f"Unknown preset ID: {preset_id}, using default")
        return ExcelFormattingPreset.from_dict(self.PRESETS["default"])

    def get_available_presets(self) -> Dict[str, str]:
        """Get dictionary of available presets (ID: Name)"""
        return {preset_id: data["name"] for preset_id, data in self.PRESETS.items()}

    def set_preset(self, preset_id: str):
        """Set current formatting preset"""
        self.current_preset = self.get_preset(preset_id)
        logger.info(f"Set Excel formatting preset to: {preset_id}")

    def format_workbook(self, file_path: str, df: pd.DataFrame, sheet_name: str = "Clearance",
                      unit: UnitType = UnitType.MIL, title: str = None) -> bool:
        """
        Apply formatting to an Excel workbook
        
        Args:
            file_path: Path to the Excel file (will be created/overwritten)
            df: DataFrame to write
            sheet_name: Name of the sheet
            unit: Unit type for values
            title: Optional title for the sheet
            
        Returns:
            True if successful, False otherwise
        """
        try:
            # First, write the DataFrame to Excel
            writer = pd.ExcelWriter(file_path, engine='openpyxl')
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Apply base formatting
            self._apply_base_formatting(worksheet, df)
            
            # Format headers
            self._format_headers(worksheet, df)
            
            # Format data cells
            self._format_data_cells(worksheet, df, unit)
            
            # Apply conditional formatting
            self._apply_conditional_formatting(worksheet, df)
            
            # Add title if provided
            if title:
                self._add_title(worksheet, title, df.shape[1])
            
            # Add summary if enabled
            if self.current_preset.include_summary:
                self._add_summary(workbook, df, unit)
            
            # Add statistics if enabled
            if self.current_preset.include_statistics:
                self._add_statistics(workbook, df, unit)
            
            # Add chart if enabled
            if self.current_preset.chart_type:
                self._add_chart(workbook, worksheet, df, self.current_preset.chart_type)
            
            # Save the workbook
            writer.close()
            
            logger.info(f"Successfully formatted Excel file: {file_path}")
            return True
            
        except Exception as e:
            error_msg = f"Error formatting Excel file: {str(e)}"
            logger.error(error_msg)
            return False

    def _apply_base_formatting(self, worksheet, df):
        """Apply base formatting to worksheet"""
        # Set column widths
        for i, column in enumerate(df.columns):
            column_width = max(len(str(column)), df[column].astype(str).str.len().max())
            column_width = min(column_width + 2, 30)  # Add padding, cap at 30
            worksheet.column_dimensions[get_column_letter(i + 1)].width = column_width
        
        # Freeze panes if enabled
        if self.current_preset.freeze_panes:
            worksheet.freeze_panes = "B2"
        
        # Apply auto filter if enabled
        if self.current_preset.auto_filter:
            worksheet.auto_filter.ref = worksheet.dimensions

    def _format_headers(self, worksheet, df):
        """Format header row"""
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        for col_num, column in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Add thin border
            cell.border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

    def _format_data_cells(self, worksheet, df, unit):
        """Format data cells based on content"""
        # Define styles
        number_format = '#,##0.00' if unit == UnitType.MM or unit == UnitType.INCH else '#,##0'
        net_class_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        data_font = Font(name='Calibri', size=11)
        thin_border = Border(
            left=Side(style='thin', color='D4D4D4'),
            right=Side(style='thin', color='D4D4D4'),
            top=Side(style='thin', color='D4D4D4'),
            bottom=Side(style='thin', color='D4D4D4')
        )
        
        # Format net class column (first column)
        for row_num in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_num, column=1)
            cell.font = Font(bold=True, size=11)
            cell.fill = net_class_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.border = thin_border
        
        # Format data cells
        for row_num in range(2, len(df) + 2):
            for col_num in range(2, len(df.columns) + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Set number format for numeric cells
                value = worksheet.cell(row=row_num, column=col_num).value
                if isinstance(value, (int, float)) and not pd.isna(value):
                    cell.number_format = number_format

    def _apply_conditional_formatting(self, worksheet, df):
        """Apply conditional formatting based on data values"""
        # Skip first column (net class names)
        data_range = f"B2:{get_column_letter(len(df.columns))}{len(df) + 1}"
        
        # Add color scale conditional formatting
        color_scale = ColorScaleRule(
            start_type='min', start_color='63BE7B',  # Green
            mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Yellow
            end_type='max', end_color='F8696B'  # Red
        )
        worksheet.conditional_formatting.add(data_range, color_scale)

    def _add_title(self, worksheet, title, num_columns):
        """Add title to worksheet"""
        # Insert a row at the top
        worksheet.insert_rows(1)
        
        # Add title cell
        title_cell = worksheet.cell(row=1, column=1)
        title_cell.value = title
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Merge cells for title
        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_columns)
        
        # Adjust row height
        worksheet.row_dimensions[1].height = 24

    def _add_summary(self, workbook, df, unit):
        """Add summary sheet with key information"""
        # Create summary sheet
        summary_sheet = workbook.create_sheet(title="Summary")
        
        # Add title
        summary_sheet.cell(row=1, column=1).value = "Clearance Rules Summary"
        summary_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        summary_sheet.merge_cells('A1:D1')
        
        # Add basic information
        summary_sheet.cell(row=3, column=1).value = "Number of Net Classes:"
        summary_sheet.cell(row=3, column=2).value = len(df) - 1  # Subtract header
        
        summary_sheet.cell(row=4, column=1).value = "Unit:"
        summary_sheet.cell(row=4, column=2).value = unit.value
        
        summary_sheet.cell(row=5, column=1).value = "Minimum Clearance:"
        # Calculate minimum clearance from data cells (skipping first column)
        numeric_data = df.iloc[:, 1:].select_dtypes(include=['number'])
        min_value = numeric_data.min().min() if not numeric_data.empty else "N/A"
        summary_sheet.cell(row=5, column=2).value = min_value
        
        summary_sheet.cell(row=6, column=1).value = "Maximum Clearance:"
        max_value = numeric_data.max().max() if not numeric_data.empty else "N/A"
        summary_sheet.cell(row=6, column=2).value = max_value
        
        summary_sheet.cell(row=7, column=1).value = "Average Clearance:"
        avg_value = numeric_data.mean().mean() if not numeric_data.empty else "N/A"
        summary_sheet.cell(row=7, column=2).value = round(avg_value, 2) if isinstance(avg_value, (int, float)) else avg_value
        
        # Format summary sheet
        for row in range(3, 8):
            summary_sheet.cell(row=row, column=1).font = Font(bold=True)
        
        # Set column widths
        summary_sheet.column_dimensions['A'].width = 25
        summary_sheet.column_dimensions['B'].width = 15

    def _add_statistics(self, workbook, df, unit):
        """Add statistics sheet with detailed analysis"""
        # Create statistics sheet
        stats_sheet = workbook.create_sheet(title="Statistics")
        
        # Add title
        stats_sheet.cell(row=1, column=1).value = "Clearance Rules Statistics"
        stats_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        stats_sheet.merge_cells('A1:D1')
        
        # Extract net class names (excluding the first column header)
        net_classes = df.iloc[:, 0].tolist()
        
        # Create statistics table
        stats_sheet.cell(row=3, column=1).value = "Net Class"
        stats_sheet.cell(row=3, column=2).value = "Min Clearance"
        stats_sheet.cell(row=3, column=3).value = "Max Clearance"
        stats_sheet.cell(row=3, column=4).value = "Avg Clearance"
        
        # Format header
        for col in range(1, 5):
            stats_sheet.cell(row=3, column=col).font = Font(bold=True)
            stats_sheet.cell(row=3, column=col).fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            stats_sheet.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")
        
        # Calculate statistics for each net class (row)
        for i, net_class in enumerate(net_classes):
            row_num = i + 4
            stats_sheet.cell(row=row_num, column=1).value = net_class
            
            # Get row data (excluding first column which is the net class name)
            row_data = df.iloc[i, 1:].select_dtypes(include=['number'])
            
            # Calculate statistics
            if not row_data.empty:
                stats_sheet.cell(row=row_num, column=2).value = row_data.min()
                stats_sheet.cell(row=row_num, column=3).value = row_data.max()
                stats_sheet.cell(row=row_num, column=4).value = round(row_data.mean(), 2)
            else:
                stats_sheet.cell(row=row_num, column=2).value = "N/A"
                stats_sheet.cell(row=row_num, column=3).value = "N/A"
                stats_sheet.cell(row=row_num, column=4).value = "N/A"
        
        # Set column widths
        stats_sheet.column_dimensions['A'].width = 25
        stats_sheet.column_dimensions['B'].width = 15
        stats_sheet.column_dimensions['C'].width = 15
        stats_sheet.column_dimensions['D'].width = 15
        
        # Apply auto filter
        last_row = len(net_classes) + 3
        stats_sheet.auto_filter.ref = f"A3:D{last_row}"

    def _add_chart(self, workbook, worksheet, df, chart_type):
        """Add chart to visualize clearance data"""
        if chart_type == "bar":
            self._add_bar_chart(workbook, df)
        elif chart_type == "heatmap":
            self._add_heatmap(workbook, worksheet, df)
        # Add other chart types as needed

    def _add_bar_chart(self, workbook, df):
        """Add bar chart of average clearances"""
        # Create chart sheet
        chart_sheet = workbook.create_sheet(title="Charts")
        
        # Add title
        chart_sheet.cell(row=1, column=1).value = "Clearance Visualization"
        chart_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Create helper table for chart data
        chart_sheet.cell(row=3, column=1).value = "Net Class"
        chart_sheet.cell(row=3, column=2).value = "Average Clearance"
        
        # Format header
        for col in range(1, 3):
            chart_sheet.cell(row=3, column=col).font = Font(bold=True)
        
        # Calculate average clearance for each net class
        net_classes = df.iloc[:, 0].tolist()
        for i, net_class in enumerate(net_classes):
            row_num = i + 4
            chart_sheet.cell(row=row_num, column=1).value = net_class
            
            # Get row data (excluding first column)
            row_data = df.iloc[i, 1:].select_dtypes(include=['number'])
            
            # Calculate average
            if not row_data.empty:
                chart_sheet.cell(row=row_num, column=2).value = row_data.mean()
            else:
                chart_sheet.cell(row=row_num, column=2).value = 0
        
        # Create bar chart
        chart = BarChart()
        chart.title = "Average Clearance by Net Class"
        chart.x_axis.title = "Net Class"
        chart.y_axis.title = "Clearance"
        
        # Define data ranges
        last_row = len(net_classes) + 3
        categories = Reference(chart_sheet, min_col=1, min_row=4, max_row=last_row)
        data = Reference(chart_sheet, min_col=2, min_row=3, max_row=last_row)
        
        # Add data to chart
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        
        # Add chart to sheet
        chart_sheet.add_chart(chart, "D3")
        
        # Set column widths
        chart_sheet.column_dimensions['A'].width = 25
        chart_sheet.column_dimensions['B'].width = 15

    def _add_heatmap(self, workbook, worksheet, df):
        """Create a heatmap visualization using conditional formatting"""
        # Ensure worksheet has conditional formatting applied
        data_range = f"B2:{get_column_letter(len(df.columns))}{len(df) + 1}"
        
        # Create more pronounced color scale for heatmap
        color_scale = ColorScaleRule(
            start_type='min', start_color='63BE7B',  # Green
            mid_type='percentile', mid_value=50, mid_color='FFEB84',  # Yellow
            end_type='max', end_color='F8696B'  # Red
        )
        worksheet.conditional_formatting.add(data_range, color_scale)
        
        # Create a legend sheet
        legend_sheet = workbook.create_sheet(title="Legend")
        
        # Add title
        legend_sheet.cell(row=1, column=1).value = "Heatmap Legend"
        legend_sheet.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Create legend
        legend_sheet.cell(row=3, column=1).value = "Color Scale Interpretation:"
        legend_sheet.cell(row=3, column=1).font = Font(bold=True)
        
        legend_sheet.cell(row=5, column=1).value = "Minimum Value (Green)"
        legend_sheet.cell(row=5, column=2).value = "Example"
        legend_sheet.cell(row=5, column=2).fill = PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid")
        
        legend_sheet.cell(row=6, column=1).value = "Medium Value (Yellow)"
        legend_sheet.cell(row=6, column=2).value = "Example"
        legend_sheet.cell(row=6, column=2).fill = PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid")
        
        legend_sheet.cell(row=7, column=1).value = "Maximum Value (Red)"
        legend_sheet.cell(row=7, column=2).value = "Example"
        legend_sheet.cell(row=7, column=2).fill = PatternFill(start_color="F8696B", end_color="F8696B", fill_type="solid")
        
        # Add explanation
        legend_sheet.cell(row=9, column=1).value = "Interpretation:"
        legend_sheet.cell(row=9, column=1).font = Font(bold=True)
        
        legend_sheet.cell(row=10, column=1).value = "This heatmap visualizes clearance values between net classes."
        legend_sheet.cell(row=11, column=1).value = "Darker green indicates smaller clearance values."
        legend_sheet.cell(row=12, column=1).value = "Darker red indicates larger clearance values."
        
        # Set column widths
        legend_sheet.column_dimensions['A'].width = 30
        legend_sheet.column_dimensions['B'].width = 15
